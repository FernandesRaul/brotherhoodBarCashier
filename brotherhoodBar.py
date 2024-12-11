import os
import tkinter as tk
from datetime import datetime
from tkinter import ttk, messagebox, simpledialog
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook

# Gravação dos arquivos
DADOS_FILE = "dados.xlsx"

# Controle de autenticação do usuário
usuario_autenticado = False


# Inicializar os arquivos necessários
def inicializar_arquivos():
    if not os.path.exists(DADOS_FILE):
        wb = Workbook()
        estoque_ws = wb.create_sheet("Estoque")
        vendas_ws = wb.create_sheet("Vendas")

        # Configuração da aba "estoque"
        estoque_ws.append(["ID", "Produto", "Preço", "Quantidade"])

        # Configuração da aba "vendas"
        vendas_ws.append(["ID Venda", "Produto", "Quantidade", "Preço Total", "Forma de Pagamento", "Data"])

        wb.save(DADOS_FILE)


# Tela de login
def tela_login():
    def validar_login():
        global usuario_autenticado
        usuario = login_var.get()
        senha = senha_var.get()
        if usuario == "Sim" and senha == "sim":
            usuario_autenticado = True
            atualizar_menu()
            exibir_conteudo(tela_inicial)
        else:
            messagebox.showerror("Erro de Login", "Usuário ou senha inválidos")

    # Limpar a tela de conteúdo
    for widget in frame_conteudo.winfo_children():
        widget.destroy()

    # Título da tela de login
    tk.Label(frame_conteudo, text=" - BROTHERHOOD MC - \n Controle de Caixa e Estoque", font=("Arial", 35, "bold"), fg="red", bg="black").pack(pady=120)

    # Campos de login
    login_var = tk.StringVar()
    tk.Label(frame_conteudo, text="Usuário", fg="white", bg="black", font=("Arial", 16)).pack(pady=5)
    tk.Entry(frame_conteudo, textvariable=login_var, font=("Arial", 14)).pack(pady=5)

    senha_var = tk.StringVar()
    tk.Label(frame_conteudo, text="Senha", fg="white", bg="black", font=("Arial", 16)).pack(pady=5)
    tk.Entry(frame_conteudo, textvariable=senha_var, show="*", font=("Arial", 14)).pack(pady=5)

    # Botão "Entrar"
    tk.Button(frame_conteudo, text="Entrar", command=validar_login, bg="#e60000", fg="white", font=("Arial", 16),
              width=10, height=1).pack(pady=20)


# Função para limpar e exibir novos conteúdos pós-login
def exibir_conteudo(funcao):
    if not usuario_autenticado and funcao != tela_login:
        messagebox.showwarning("Acesso Negado", "Você precisa fazer login para acessar esta funcionalidade.")
        return
    for widget in frame_conteudo.winfo_children():
        widget.destroy()
    funcao()


# Atualizar a visibilidade dos botões do menu pós-login
def atualizar_menu():
    for widget in frame_menu.winfo_children():
        widget.destroy()

    if not usuario_autenticado:
        tk.Button(frame_menu, text="Login", command=lambda: exibir_conteudo(tela_login), bg="#e60000", fg="white",
                  font=("Arial", 12, "bold")).pack(fill="x", pady=10, padx=10)
    else:
        botoes = [
            {"text": "Tela Inicial", "command": lambda: exibir_conteudo(tela_inicial)},
            {"text": "Caixa", "command": lambda: exibir_conteudo(caixa)},
            {"text": "Estoque", "command": lambda: exibir_conteudo(estoque)},
            {"text": "Adicionar Estoque", "command": lambda: exibir_conteudo(adicionar_estoque)},
            {"text": "Auditoria", "command": lambda: exibir_conteudo(auditoria)}
        ]
        for botao in botoes:
            tk.Button(frame_menu, text=botao["text"], command=botao["command"], bg="#e60000", fg="white",
                      font=("Arial", 12, "bold")).pack(fill="x", pady=15, padx=10)


# Função Tela Inicial
def tela_inicial():
    if background_image:
        background_label = tk.Label(frame_conteudo, image=background_image, bg="black")
        background_label.place(relwidth=1, relheight=1)
    tk.Label(frame_conteudo, text="- BROTHERHOOD MC -\n WE ARE DISEASE UPON THE WORLD!", font=("Arial", 20, "bold"),
             fg="red", bg="black").pack(pady=50)


# Função Caixa
def caixa():
    tk.Label(frame_conteudo, text="Caixa", font=("Arial", 35, "bold"), fg="red", bg="black").pack(pady=50)

    wb = load_workbook(DADOS_FILE)
    estoque_ws = wb["Estoque"]

    produtos = {}
    for row in estoque_ws.iter_rows(min_row=2, values_only=True):
        produto, preco, quantidade = row[1], row[2], row[3]
        produtos[produto] = {"preco": preco, "quantidade": quantidade}

    selected_produto = tk.StringVar()
    selected_quantidade = tk.IntVar(value=1)

    def atualizar_valor_a_ser_pago():
        produto_selecionado = selected_produto.get()
        quantidade_selecionada = selected_quantidade.get()

        preco_produto = produtos.get(produto_selecionado, {}).get("preco", 0)
        total = preco_produto * quantidade_selecionada

        valor_a_ser_pago_var.set(f"R${total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

    def confirmar_venda():
        produto_selecionado = selected_produto.get()
        quantidade_selecionada = selected_quantidade.get()
        preco_produto = produtos.get(produto_selecionado, {}).get("preco", 0)
        quantidade_produto = produtos.get(produto_selecionado, {}).get("quantidade", 0)

        if quantidade_produto >= quantidade_selecionada:
            produtos[produto_selecionado]["quantidade"] -= quantidade_selecionada

            for row1 in estoque_ws.iter_rows(min_row=2):
                if row1[1].value == produto_selecionado:
                    row1[3].value -= quantidade_selecionada
                    break

            total_produto = preco_produto * quantidade_selecionada
            vendas_ws = wb["Vendas"]
            vendas_ws.append(
                [vendas_ws.max_row, produto_selecionado, quantidade_selecionada, total_produto, pagamento_var.get(),
                 datetime.now()])

            wb.save(DADOS_FILE)

            if pagamento_var.get() == "Dinheiro":
                valor_recebido = simpledialog.askfloat("Valor Recebido", "Digite o valor recebido:", parent=root,
                                                       minvalue=0)
                troco = valor_recebido - total_produto if valor_recebido else 0
                messagebox.showinfo("Venda Efetuada",
                                    f"Venda efetuada com sucesso!\nTroco: R${troco:,.2f}".replace(",", "X").replace(".",
                                                                                                                    ",").replace(
                                        "X", "."))
            else:
                messagebox.showinfo("Venda Efetuada", "Venda efetuada com sucesso!")
        else:
            messagebox.showwarning("Sem Estoque", f"Estamos sem estoque suficiente de {produto_selecionado}.")

    # Layout Caixa
    tk.Label(frame_conteudo, text="SELECIONE O PRODUTO", fg="white", bg="black", font=("Arial", 14)).pack(pady=15)
    produto_dropdown = ttk.Combobox(frame_conteudo, textvariable=selected_produto, values=list(produtos.keys()),
                                    font=("Arial", 14))
    produto_dropdown.pack(pady=15)

    tk.Label(frame_conteudo, text="QUANTIDADE", fg="white", bg="black", font=("Arial", 14)).pack(pady=15)
    quantidade_dropdown = ttk.Combobox(frame_conteudo, textvariable=selected_quantidade,
                                       values=[str(i) for i in range(1, 11)], font=("Arial", 14))
    quantidade_dropdown.pack(pady=15)

    valor_a_ser_pago_var = tk.StringVar()
    valor_a_ser_pago_var.set("R$0,00")
    tk.Label(frame_conteudo, text="VALOR A SER PAGO", fg="white", bg="black", font=("Arial", 14)).pack(pady=15)
    valor_a_ser_pago_entry = tk.Entry(frame_conteudo, textvariable=valor_a_ser_pago_var, font=("Arial", 14),
                                      state="readonly")
    valor_a_ser_pago_entry.pack(pady=15)

    selected_produto.trace("w", lambda *args: atualizar_valor_a_ser_pago())
    selected_quantidade.trace("w", lambda *args: atualizar_valor_a_ser_pago())

    tk.Label(frame_conteudo, text="FORMA DE PAGAMENTO", fg="white", bg="black", font=("Arial", 14)).pack(pady=15)
    pagamento_var = tk.StringVar(value="Dinheiro")
    pagamento_dropdown = ttk.Combobox(frame_conteudo, textvariable=pagamento_var,
                                      values=["Dinheiro", "Débito", "Crédito", "Pix"], font=("Arial", 14))
    pagamento_dropdown.pack(pady=15)

    tk.Button(frame_conteudo, text="CONFIRMAR VENDA", command=confirmar_venda, bg="#e60000", fg="white",
              font=("Arial", 14)).pack(pady=25)


# Função Estoque
def estoque():
    tk.Label(frame_conteudo, text="Estoque", font=("Arial", 35, "bold"), fg="red", bg="black").pack(pady=20)

    wb = load_workbook(DADOS_FILE)
    estoque_ws = wb["Estoque"]

    treeview = ttk.Treeview(frame_conteudo, columns=("Produto", "Preço", "Quantidade"), show="headings", height=8)
    treeview.heading("Produto", text="Produto")
    treeview.heading("Preço", text="Preço")
    treeview.heading("Quantidade", text="Quantidade")

    treeview.column("Produto", anchor="center")
    treeview.column("Preço", anchor="center")
    treeview.column("Quantidade", anchor="center")

    for row in estoque_ws.iter_rows(min_row=2, values_only=True):
        produto, preco, quantidade = row[1], row[2], row[3]
        preco_formatado = f"R${preco:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        treeview.insert("", "end", values=(produto, preco_formatado, quantidade))

    treeview.pack(fill="both", expand=True)


# Função Adicionar Estoque
def adicionar_estoque():
    tk.Label(frame_conteudo, text="Adicionar Estoque", font=("Arial", 35, "bold"), fg="red", bg="black").pack(pady=20)

    tk.Label(frame_conteudo, text="Nome do Produto", fg="white", bg="black", font=("Arial", 14)).pack(pady=10)
    nome_produto_var = tk.StringVar()
    tk.Entry(frame_conteudo, textvariable=nome_produto_var, font=("Arial", 14)).pack(padx=20, pady=5)

    tk.Label(frame_conteudo, text="Preço (R$)", fg="white", bg="black", font=("Arial", 14)).pack(pady=10)
    preco_produto_var = tk.DoubleVar()
    tk.Entry(frame_conteudo, textvariable=preco_produto_var, font=("Arial", 14)).pack(padx=20, pady=5)

    tk.Label(frame_conteudo, text="Quantidade", fg="white", bg="black", font=("Arial", 14)).pack(pady=10)
    quantidade_produto_var = tk.IntVar()
    tk.Entry(frame_conteudo, textvariable=quantidade_produto_var, font=("Arial", 14)).pack(padx=20, pady=5)

    def adicionar():
        nome_produto = nome_produto_var.get()
        preco_produto = preco_produto_var.get()
        quantidade_produto = quantidade_produto_var.get()

        if nome_produto and preco_produto > 0 and quantidade_produto > 0:
            wb = load_workbook(DADOS_FILE)
            estoque_ws = wb["Estoque"]
            novo_id = estoque_ws.max_row
            estoque_ws.append([novo_id, nome_produto, preco_produto, quantidade_produto])
            wb.save(DADOS_FILE)
            messagebox.showinfo("Produto Adicionado", "Produto adicionado ao estoque com sucesso!")
            exibir_conteudo(estoque)
        else:
            messagebox.showwarning("Dados Inválidos", "Preencha todos os campos corretamente.")

    tk.Button(frame_conteudo, text="Adicionar Produto", command=adicionar, bg="#e60000", fg="white",
              font=("Arial", 14)).pack(pady=20)


# Função Auditoria
def auditoria():
    tk.Label(frame_conteudo, text="Auditoria de Caixa", font=("Arial", 35, "bold"), fg="red", bg="black").pack(
        pady=20)

    tk.Label(frame_conteudo, text="Selecione o Mês", fg="white", bg="black", font=("Arial", 14)).pack(pady=10)
    mes_var = tk.StringVar()
    mes_dropdown = ttk.Combobox(frame_conteudo, textvariable=mes_var, values=[
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro",
        "Dezembro"
    ], font=("Arial", 14))
    mes_dropdown.pack(pady=10)

    tk.Label(frame_conteudo, text="Selecione o Ano", fg="white", bg="black", font=("Arial", 14)).pack(pady=10)
    ano_var = tk.StringVar()
    ano_dropdown = ttk.Combobox(frame_conteudo, textvariable=ano_var, values=[str(i) for i in range(2020, 2031)],
                                font=("Arial", 14))
    ano_dropdown.pack(pady=10)

    def gerar_balanco():
        mes = mes_var.get()
        ano = ano_var.get()
        if mes and ano:
            balanco = calcular_balanco(mes, ano)
            messagebox.showinfo("Balanço", f"Balanço para {mes}/{ano}:\n\n{balanco}")
        else:
            messagebox.showwarning("Erro", "Por favor, selecione o mês e o ano.")

    tk.Button(frame_conteudo, text="Gerar Balanço", command=gerar_balanco, bg="#e60000", fg="white",
              font=("Arial", 14)).pack(pady=20)


# Função de auditoria de caixa (simplificada)
def calcular_balanco(mes, ano):
    wb = load_workbook(DADOS_FILE)
    vendas_ws = wb["Vendas"]
    total_vendas = 0

    for row in vendas_ws.iter_rows(min_row=2, values_only=True):
        data_venda = row[5]
        if data_venda and data_venda.month == datetime.strptime(mes, "%B").month and data_venda.year == int(ano):
            total_vendas += row[3]  # somando o preço total

    return f"Total de Vendas: R${total_vendas:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# Interface Principal
root = tk.Tk()
root.state("zoomed")
root.title("Gestão de Bar")
root.geometry("1200x800")
root.configure(bg="black")

# Estilo
style = ttk.Style(root)
style.theme_use("clam")
style.configure("Treeview", background="#333", fieldbackground="#333", foreground="white")
style.configure("Treeview.Heading", background="#444", foreground="white", font=("Arial", 12, "bold"))

# Layout
frame_menu = tk.Frame(root, bg="#111", width=250)
frame_menu.pack(side="left", fill="y")

frame_conteudo = tk.Frame(root, bg="black")
frame_conteudo.pack(side="right", expand=True, fill="both")

# Imagem de Fundo com Tratamento de Erros
try:
    background_image = ImageTk.PhotoImage(Image.open("background.jpeg").resize((800, 800)))
except Exception as e:
    print(f"Erro ao carregar a imagem de fundo: {e}")
    background_image = None

# Inicializar arquivos e menu
inicializar_arquivos()
atualizar_menu()
tela_login()
root.mainloop()
